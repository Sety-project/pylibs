
import openpyxl
import json
import os
from pathlib import Path
import datetime

def excel_to_dict(excel_path, sheet_name, headers=[]):
    def _index_to_col(index):
        q = int((index-1)/26)
        rest = (index - q*26)
        if q == 0:
            return chr(ord('@') + index)
        else:
            return chr(ord('@') + q) + chr(ord('@') + rest)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.get_sheet_by_name(sheet_name)

    # List all headers
    if not headers:
        for cell in ws[1]:
            headers.append(cell.value)

    result_dict = {}
    for row in range(2, ws.max_row +1):
        line = dict()
        for header in headers:
            cell_value = ws[_index_to_col(headers.index(header)+1) + str(row)].value
            if isinstance(cell_value, str):
                cell_value = cell_value.encode('utf-8').decode('ascii', 'ignore')
                cell_value = cell_value.strip()
            elif type(cell_value) is datetime:
                cell_value = str(cell_value)
            elif cell_value is None:
                cell_value = ''
            line[header] = cell_value
        result_dict[line['name']] = {k:v for k,v in line.items() if k != 'name'}
    return result_dict

def convert_api_param_to_json():
    dict_to_dump = excel_to_dict('/home/vic/pylibs/SystematicCeFi/DerivativeArbitrage/Runtime/configs/static_params.xlsx', 'api', ['exchange', 'key', 'comment'])
    with open(os.path.join(Path.home(), '.cache', 'setyvault', 'api_keys.json'), 'w') as fp:
        json.dump(dict_to_dump, fp, indent=4)

def convert_ss_to_json():
    dict_to_dump = excel_to_dict('/home/vic/mktdata/universe2.xlsx', 'max')
    with open(os.path.join(Path.home(), 'mktdata', 'universe.json'), 'w') as fp:
        json.dump(dict_to_dump, fp, indent=4, default=str)
    print("Exiting gracefully")